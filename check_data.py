import openpyxl

wb = openpyxl.load_workbook('mai_js.xlsx')
ws = wb['Sheet1']

print(f'Sheet: Sheet1')
print(f'Max row: {ws.max_row}')
print(f'Max column: {ws.max_column}')
print()

# Check row 2
print('Checking row 2:')
print(f'  I2 (fc_code_ref): {ws["I2"].value}')
print(f'  AA2 (check): {ws["AA2"].value}')
print()

# Check row 100
print('Checking row 100:')
print(f'  I100 (fc_code_ref): {ws["I100"].value}')
print(f'  AA100 (check): {ws["AA100"].value}')
print()

# Check row 990
print('Checking row 990:')
print(f'  I990 (fc_code_ref): {ws["I990"].value}')
print(f'  AA990 (check): {ws["AA990"].value}')
print()

# Count non-empty cells in column AA (check)
count_aa = 0
count_i = 0
for row in range(2, ws.max_row + 1):
    if ws[f'AA{row}'].value:
        count_aa += 1
    if ws[f'I{row}'].value:
        count_i += 1

print(f'Total rows with data in column AA (check): {count_aa}')
print(f'Total rows with data in column I (fc_code_ref): {count_i}')
