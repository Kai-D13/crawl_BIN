import openpyxl

wb = openpyxl.load_workbook('D:/WHM_BIN/just_test.xlsx')
ws = wb.active

print('Total rows:', ws.max_row)
print('Total columns:', ws.max_column)
print('\nColumns G, H, I:')
print('G (RFID), H (FC_CODE_REF), I (LINK_INTERNAL)\n')

for row in range(1, min(12, ws.max_row + 1)):
    g_val = ws[f'G{row}'].value
    h_val = ws[f'H{row}'].value
    i_val = ws[f'I{row}'].value
    i_display = str(i_val)[:80] + '...' if i_val and len(str(i_val)) > 80 else i_val
    print(f'Row {row}: G={g_val}, H={h_val}')
    print(f'        I={i_display}')
    print()
