import openpyxl

wb = openpyxl.load_workbook('D:/WHM_BIN/just_test_1.xlsx')
ws = wb.active

print('=== KIỂM TRA CẤU TRÚC FILE EXCEL ===\n')
print(f'Sheet name: {ws.title}')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}\n')

# Tìm header
print('Headers (Row 1):')
for col in range(1, min(30, ws.max_column + 1)):
    cell = ws.cell(1, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        print(f'  {col_letter} (col {col}): {cell.value}')

# Tìm cột reference_code_of_so
print('\n=== TÌM CỘT QUAN TRỌNG ===')
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    if header and 'reference_code_of_so' in str(header).lower():
        print(f'✓ reference_code_of_so: Cột {col_letter} (col {col})')
        # Show some data
        print(f'  Ví dụ dữ liệu:')
        for row in range(2, min(6, ws.max_row + 1)):
            val = ws.cell(row, col).value
            print(f'    Row {row}: {val}')
        break

# Kiểm tra cột I (fc_code_ref)
print(f'\n✓ fc_code_ref: Cột I (col 9)')
print(f'  Ví dụ dữ liệu:')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 9).value
    print(f'    Row {row}: {val}')

# Kiểm tra cột AA (nơi sẽ ghi kết quả)
print(f'\n✓ Cột AA (col 27) - Nơi ghi kết quả:')
print(f'  Header hiện tại: {ws.cell(1, 27).value}')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 27).value
    print(f'    Row {row}: {val}')
