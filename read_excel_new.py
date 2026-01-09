import openpyxl

wb = openpyxl.load_workbook('D:/WHM_BIN/just_test_1.xlsx')
ws = wb.active

print('Total rows:', ws.max_row)
print('Total columns:', ws.max_column)
print('\nColumns I (fc_code_ref), J (link_internal):\n')

for row in range(1, min(12, ws.max_row + 1)):
    i_val = ws[f'I{row}'].value
    j_val = ws[f'J{row}'].value
    j_display = str(j_val)[:100] + '...' if j_val and len(str(j_val)) > 100 else j_val
    print(f'Row {row}:')
    print(f'  I (fc_code_ref) = {i_val}')
    print(f'  J (link_internal) = {j_display}')
    print()
