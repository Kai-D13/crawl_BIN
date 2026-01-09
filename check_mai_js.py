import openpyxl

wb = openpyxl.load_workbook('D:/WHM_BIN/mai_js.xlsx')

print('=== THÔNG TIN FILE EXCEL MỚI: mai_js.xlsx ===\n')
print(f'Danh sách sheets:')
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'  {i}. "{sheet_name}"')

ws = wb.active
print(f'\nActive sheet: "{ws.title}"')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}\n')

# Kiểm tra headers
print('=== HEADERS (Row 1) ===')
for col in range(1, min(30, ws.max_column + 1)):
    cell = ws.cell(1, col)
    col_letter = openpyxl.utils.get_column_letter(col)
    if cell.value:
        print(f'{col_letter} (col {col:2d}): {cell.value}')

# Kiểm tra các cột quan trọng
print('\n=== KIỂM TRA CÁC CỘT QUAN TRỌNG ===')

# Cột I (fc_code_ref)
print(f'\nCột I (fc_code_ref):')
print(f'  Header: {ws.cell(1, 9).value}')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 9).value
    print(f'  Row {row}: {val}')

# Cột J (link_internal)
print(f'\nCột J (link_internal):')
print(f'  Header: {ws.cell(1, 10).value}')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 10).value
    val_display = str(val)[:80] + '...' if val and len(str(val)) > 80 else val
    print(f'  Row {row}: {val_display}')

# Cột K (reference_code_of_so)
print(f'\nCột K (reference_code_of_so):')
print(f'  Header: {ws.cell(1, 11).value}')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 11).value
    print(f'  Row {row}: {val}')

# Cột AA (check)
print(f'\nCột AA (check - col 27):')
print(f'  Header: {ws.cell(1, 27).value}')
for row in range(2, min(6, ws.max_row + 1)):
    val = ws.cell(row, 27).value
    print(f'  Row {row}: {val}')
